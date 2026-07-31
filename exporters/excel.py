from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from storage.repository import Repository


class ExcelExporter:
    def __init__(self, repository: Repository) -> None:
        self.repository = repository

    def export(self, output_path: Path | str) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        default = workbook.active
        workbook.remove(default)

        self._write_sheet(workbook, "总览", self.repository.fetch_overview())
        self._write_sheet(workbook, "招聘信息", self.repository.fetch_all("recruitments"))
        self._write_sheet(workbook, "面经", self.repository.fetch_all("interviews"))
        self._write_sheet(workbook, "面试轮次", self.repository.fetch_all("interview_rounds"))
        self._write_sheet(workbook, "Offer", self.repository.fetch_all("offers"))
        self._write_sheet(
            workbook,
            "待遇与工作体验",
            self.repository.fetch_all("work_conditions"),
        )
        self._write_sheet(workbook, "待人工确认", self.repository.fetch_needs_review())

        workbook.save(path)
        return path

    def _write_sheet(
        self, workbook: Workbook, sheet_name: str, rows: list[dict[str, Any]]
    ) -> None:
        worksheet = workbook.create_sheet(sheet_name)
        rows = [_display_row(row) for row in rows]
        headers = list(rows[0].keys()) if rows else ["empty"]

        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")

        for row in rows:
            worksheet.append([row.get(header) for header in headers])

        worksheet.freeze_panes = "A2"
        for idx, header in enumerate(headers, start=1):
            max_length = max(
                [len(str(header))]
                + [
                    len(str(worksheet.cell(row=row_idx, column=idx).value or ""))
                    for row_idx in range(2, worksheet.max_row + 1)
                ]
            )
            width = min(max(max_length + 2, 10), 48)
            worksheet.column_dimensions[get_column_letter(idx)].width = width


def _display_row(row: dict[str, Any]) -> dict[str, Any]:
    display: dict[str, Any] = {}
    for key, value in row.items():
        output_key = key.removesuffix("_json")
        if key.endswith("_json"):
            display[output_key] = _json_to_cell(value)
        elif key == "needs_review":
            display[key] = "是" if value else "否"
        elif key in {"accepted", "self_intro"} and value is not None:
            display[key] = "是" if value else "否"
        else:
            display[key] = value
    return display


def _json_to_cell(value: str | None) -> str:
    if not value:
        return ""
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return value
    if parsed is None:
        return ""
    if isinstance(parsed, list):
        return "；".join(str(item) for item in parsed)
    return json.dumps(parsed, ensure_ascii=False)

