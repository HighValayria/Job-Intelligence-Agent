from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from exporters.excel import ExcelExporter
from scheduler.runner import PipelineRunner
from storage.repository import Repository


def test_excel_export_contains_required_sheets(tmp_path: Path) -> None:
    db_path = tmp_path / "job.sqlite3"
    excel_path = tmp_path / "job_intelligence.xlsx"

    PipelineRunner(db_path=db_path, excel_path=excel_path).run()

    with Repository(db_path) as repository:
        ExcelExporter(repository).export(excel_path)

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == [
        "总览",
        "招聘信息",
        "面经",
        "面试轮次",
        "Offer",
        "待遇与工作体验",
        "待人工确认",
    ]
    assert workbook["总览"].max_row == 5
    assert workbook["Offer"].max_row == 2

